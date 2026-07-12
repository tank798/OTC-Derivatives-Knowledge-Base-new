from __future__ import annotations

import argparse
import csv
import hashlib
import json
from pathlib import Path
import re
from urllib.parse import urlparse

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
METADATA_PATH = ROOT / "data/metadata/regulations.jsonl"
INDEX_METADATA_PATH = ROOT / "data/index/document_metadata.jsonl"
CORPUS_PATH = ROOT / "data/index/corpus.jsonl"
MANIFEST_PATH = ROOT / "data/index/manifest.json"
VECTOR_METADATA_PATH = ROOT / "data/index/vector_metadata.json"


def load_jsonl(path: Path) -> list[dict]:
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def write_jsonl(path: Path, rows: list[dict]) -> None:
    path.write_text("".join(json.dumps(row, ensure_ascii=False) + "\n" for row in rows), encoding="utf-8")


def normalize(value: str) -> str:
    return re.sub(r"[\s《》【】\[\]（）()，,。:：;；'\"“”‘’]", "", str(value or "")).lower()


def valid_url(value: str) -> bool:
    parsed = urlparse(value.strip())
    return parsed.scheme in {"http", "https"} and bool(parsed.netloc)


def sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def main() -> int:
    parser = argparse.ArgumentParser(description="将人工补充的法规官网URL同步回知识库元数据")
    parser.add_argument("--xlsx", type=Path, default=ROOT / "outputs/missing_regulation_urls.xlsx")
    parser.add_argument("--conflicts", type=Path, default=ROOT / "outputs/url_update_conflicts.csv")
    args = parser.parse_args()

    workbook = load_workbook(args.xlsx, data_only=True)
    sheet = workbook.active
    headers = {str(cell.value).strip(): index for index, cell in enumerate(sheet[1]) if cell.value}
    if "法规名称" not in headers or "URL" not in headers:
        raise SystemExit("Excel必须包含“法规名称”和“URL”两列")

    metadata = load_jsonl(METADATA_PATH)
    by_title: dict[str, list[int]] = {}
    by_title_number: dict[tuple[str, str], list[int]] = {}
    for index, row in enumerate(metadata):
        title = normalize(row.get("document_title", ""))
        number = normalize(row.get("document_number", ""))
        by_title.setdefault(title, []).append(index)
        if number:
            by_title_number.setdefault((title, number), []).append(index)

    conflicts: list[dict[str, str]] = []
    updates: dict[str, str] = {}
    for values in sheet.iter_rows(min_row=2, values_only=True):
        title = str(values[headers["法规名称"]] or "").strip()
        url = str(values[headers["URL"]] or "").strip()
        number = str(values[headers.get("文号", -1)] or "").strip() if "文号" in headers else ""
        if not title or not url:
            continue
        if not valid_url(url):
            conflicts.append({"法规名称": title, "文号": number, "URL": url, "原因": "URL格式无效"})
            continue
        candidates = by_title_number.get((normalize(title), normalize(number)), []) if number else []
        if not candidates:
            candidates = by_title.get(normalize(title), [])
        if len(candidates) != 1:
            conflicts.append({"法规名称": title, "文号": number, "URL": url, "原因": f"匹配数量为{len(candidates)}，未自动覆盖"})
            continue
        document_id = metadata[candidates[0]]["document_id"]
        updates[document_id] = url
        metadata[candidates[0]]["official_url"] = url

    write_jsonl(METADATA_PATH, metadata)
    index_metadata = load_jsonl(INDEX_METADATA_PATH)
    corpus = load_jsonl(CORPUS_PATH)
    for row in index_metadata:
        if row["document_id"] in updates:
            row["official_url"] = updates[row["document_id"]]
    for row in corpus:
        if row["document_id"] in updates:
            row["official_url"] = updates[row["document_id"]]
    write_jsonl(INDEX_METADATA_PATH, index_metadata)
    write_jsonl(CORPUS_PATH, corpus)

    args.conflicts.parent.mkdir(parents=True, exist_ok=True)
    with args.conflicts.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["法规名称", "文号", "URL", "原因"])
        writer.writeheader()
        writer.writerows(conflicts)

    manifest = json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))
    manifest["generated_at"] = __import__("datetime").datetime.now(__import__("datetime").timezone.utc).isoformat()
    manifest["corpus"]["sha256"] = sha256(CORPUS_PATH)
    manifest["corpus"]["official_url_count"] = sum(bool(row.get("official_url")) for row in index_metadata)
    MANIFEST_PATH.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    vector_metadata = json.loads(VECTOR_METADATA_PATH.read_text(encoding="utf-8"))
    vector_metadata["corpus_sha256"] = sha256(CORPUS_PATH)
    VECTOR_METADATA_PATH.write_text(json.dumps(vector_metadata, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(json.dumps({"updated": len(updates), "conflicts": len(conflicts)}, ensure_ascii=False))
    return 0 if not conflicts else 2


if __name__ == "__main__":
    raise SystemExit(main())

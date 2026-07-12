from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from models import ParsedDocument, SourceBlock
from utils.metadata import infer_metadata
from utils.text import clean_text, markdown_table


def parse_xlsx(path: Path) -> ParsedDocument:
    workbook = load_workbook(path, read_only=True, data_only=False)
    blocks: list[SourceBlock] = []
    sequence = 0
    for worksheet in workbook.worksheets:
        sequence += 1
        blocks.append(SourceBlock(f"工作表：{worksheet.title}", style="Heading 1", source_kind="sheet", block_id=f"b{sequence:05d}"))
        rows: list[list[str]] = []
        for row in worksheet.iter_rows(values_only=True):
            values = [clean_text(str(value)) if value is not None else "" for value in row]
            while values and not values[-1]:
                values.pop()
            if values and any(values):
                rows.append(values)
        if rows:
            sequence += 1
            blocks.append(SourceBlock(markdown_table(rows), style="Table", source_kind="table", block_id=f"b{sequence:05d}"))
    workbook.close()
    metadata = infer_metadata(blocks, path)
    return ParsedDocument(path, "xlsx", blocks, metadata, [])
